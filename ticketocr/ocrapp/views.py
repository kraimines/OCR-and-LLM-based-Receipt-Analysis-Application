import requests
import json
import re
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from PIL import Image
from pdf2image import convert_from_path
import pytesseract
import docling
from pathlib import Path
from django.shortcuts import render
from .forms import TicketUploadForm
from .models import ExtractionHistory, TicketHistory, AccountingEntry
from doctr.models import ocr_predictor
from doctr.io import DocumentFile
from docling.document_converter import DocumentConverter
import os
import logging
from django.conf import settings
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from io import BytesIO
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
try:
    import google.generativeai as genai
except Exception:
    genai = None
    # Use logging so the warning is emitted at WARNING level instead of printing directly.
    import logging as _logging
    _logging.warning("Warning: google.generativeai package not installed — Gemini integration disabled.")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def diagnose_system():
    """Diagnostic function to check system health"""
    issues = []
    
    # Check Ollama connection
    try:
        response = requests.get("http://localhost:11434/api/tags", timeout=5)
        if response.status_code != 200:
            issues.append("Ollama server not responding properly")
        else:
            logger.info("âœ“ Ollama server is running")
    except requests.exceptions.ConnectionError:
        issues.append("Cannot connect to Ollama server - is it running?")
    except requests.exceptions.Timeout:
        issues.append("Ollama server timeout - server may be overloaded")
    except Exception as e:
        issues.append(f"Ollama connection error: {str(e)}")
    
    # Check media directory
    try:
        media_root = getattr(settings, 'MEDIA_ROOT', None)
        if not media_root:
            issues.append("MEDIA_ROOT not configured in settings")
        elif not os.path.exists(media_root):
            issues.append(f"Media directory doesn't exist: {media_root}")
        else:
            logger.info(f"âœ“ Media directory exists: {media_root}")
    except Exception as e:
        issues.append(f"Media directory check failed: {str(e)}")
    
    # Check required Python packages
    required_packages = ['doctr', 'pytesseract', 'docling', 'PIL', 'pdf2image']
    for package in required_packages:
        try:
            __import__(package)
            logger.info(f"âœ“ {package} is available")
        except ImportError:
            issues.append(f"Missing required package: {package}")
    
    if issues:
        logger.error("System issues found:")
        for issue in issues:
            logger.error(f"  - {issue}")
    else:
        logger.info("âœ“ All system checks passed")
    
    return issues

def clean_json_response(text):
    """
    Nettoie une rÃ©ponse JSON malformÃ©e de maniÃ¨re plus robuste
    """
    if not text or not isinstance(text, str):
        return None
    
    print(f"Texte original reÃ§u: {text[:200]}...")
    
    # Supprimer les balises <think> et leur contenu
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL)
    text = re.sub(r'<.*?>', '', text)  # Supprimer toutes les balises HTML/XML
    
    # Nettoyer les espaces et sauts de ligne excessifs
    text = re.sub(r'\n+', '\n', text)
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    
    # Chercher le JSON le plus externe
    start_positions = [i for i, char in enumerate(text) if char == '{']
    end_positions = [i for i, char in enumerate(text) if char == '}']
    
    if start_positions and end_positions:
        # Essayer le JSON le plus externe d'abord
        outer_start = start_positions[0]
        outer_end = end_positions[-1]
        if outer_end > outer_start:
            json_candidate = text[outer_start:outer_end+1]
            try:
                # Nettoyer le JSON
                cleaned_json = json_candidate.strip()
                # Remplacer les guillemets simples par des guillemets doubles
                cleaned_json = re.sub(r"'([^']*)'", r'"\1"', cleaned_json)
                # Supprimer les virgules trailing
                cleaned_json = re.sub(r',(\s*[}\]])', r'\1', cleaned_json)
                
                result = json.loads(cleaned_json)
                # VÃ©rifier que c'est un objet avec les clÃ©s principales
                main_keys = ['Magasin', 'Date', 'NumeroTicket', 'Articles', 'Total']
                if isinstance(result, dict) and sum(1 for k in main_keys if k in result) >= 2:
                    print(f"âœ… JSON externe valide trouvÃ©")
                    return result
            except Exception as e:
                print(f"âŒ Erreur parsing JSON externe: {e}")
    
    # Essayer tous les JSON possibles
    best_json = None
    best_score = 0
    
    for start in start_positions:
        for end in end_positions:
            if end > start:
                json_candidate = text[start:end+1]
                try:
                    cleaned_json = json_candidate.strip()
                    cleaned_json = re.sub(r"'([^']*)'", r'"\1"', cleaned_json)
                    cleaned_json = re.sub(r',(\s*[}\]])', r'\1', cleaned_json)
                    
                    result = json.loads(cleaned_json)
                    main_keys = ['Magasin', 'Date', 'NumeroTicket', 'Articles', 'Total']
                    
                    if isinstance(result, dict):
                        score = sum(1 for k in main_keys if k in result)
                        if score >= 2 and score > best_score:
                            best_json = result
                            best_score = score
                            print(f"âœ… JSON trouvÃ© avec score {score}")
                except Exception as e:
                    print(f"âŒ Erreur parsing JSON candidat: {e}")
                    continue
    
    if best_json:
        return best_json
    
    # Dernier essai : chercher le JSON avec une regex plus permissive
    # Chercher le dernier JSON valide dans le texte
    json_pattern = r'\{(?:[^{}]|(?:\{[^{}]*\}))*\}'
    json_matches = re.findall(json_pattern, text, re.DOTALL)
    
    # Essayer le dernier JSON trouvÃ© (le plus probable d'Ãªtre complet)
    for json_match in reversed(json_matches):
        try:
            cleaned_json = json_match.strip()
            cleaned_json = re.sub(r"'([^']*)'", r'"\1"', cleaned_json)
            cleaned_json = re.sub(r',(\s*[}\]])', r'\1', cleaned_json)
            
            result = json.loads(cleaned_json)
            main_keys = ['Magasin', 'Date', 'NumeroTicket', 'Articles', 'Total']
            
            if isinstance(result, dict) and sum(1 for k in main_keys if k in result) >= 2:
                print(f"âœ… JSON trouvÃ© via regex (dernier)")
                return result
        except Exception as e:
            print(f"âŒ Erreur parsing JSON regex: {e}")
            continue
    
    # Si rien trouvÃ©, essayer de chercher juste aprÃ¨s "}" pour voir s'il y a du JSON valide
    if '}' in text:
        last_brace_pos = text.rfind('}')
        if last_brace_pos > 0:
            # Chercher le dÃ©but du JSON en remontant
            brace_count = 0
            start_pos = last_brace_pos
            for i in range(last_brace_pos, -1, -1):
                if text[i] == '}':
                    brace_count += 1
                elif text[i] == '{':
                    brace_count -= 1
                    if brace_count == 0:
                        start_pos = i
                        break
            
            if start_pos < last_brace_pos:
                json_candidate = text[start_pos:last_brace_pos + 1]
                try:
                    cleaned_json = json_candidate.strip()
                    cleaned_json = re.sub(r"'([^']*)'", r'"\1"', cleaned_json)
                    cleaned_json = re.sub(r',(\s*[}\]])', r'\1', cleaned_json)
                    
                    result = json.loads(cleaned_json)
                    main_keys = ['Magasin', 'Date', 'NumeroTicket', 'Articles', 'Total']
                    
                    if isinstance(result, dict) and sum(1 for k in main_keys if k in result) >= 2:
                        print(f"âœ… JSON trouvÃ© via recherche de braces")
                        return result
                except Exception as e:
                    print(f"âŒ Erreur parsing JSON via braces: {e}")
    
    print("âŒ Aucun JSON valide trouvÃ©")
    return None



def analyze_three_texts_with_llm_fast(ocr_results):
    """
    Version rapide avec un modÃ¨le plus lÃ©ger
    """
    docling_text = ocr_results.get("docling", "")
    tesseract_text = ocr_results.get("tesseract", "")
    doctr_text = ocr_results.get("doctr", "")
    
    prompt = f"""Tu es un assistant expert en analyse de tickets de caisse.

Voici trois extraits OCR du mÃªme ticket :

--- OCR DocLing ---
{docling_text}

--- OCR Tesseract ---
{tesseract_text}

--- OCR Doctr ---
{doctr_text}

Extrais les Ã©lÃ©ments suivants et retourne UNIQUEMENT un objet JSON valide :

{{
  "Magasin": "Nom du magasin",
  "NumeroTicket": "NumÃ©ro du ticket",
  "Date": "JJ/MM/AAAA HH:MM",
  "Articles": [
    {{ "nom": "Nom article", "prix": "Prix en DT" }},
    {{ "nom": "TIMBRE FISCAL", "prix": "0.100 DT" }}
  ],
  "Total": "Montant total en DT"
}}

RÃˆGLES IMPORTANTES :
1. Le timbre fiscal (0.100 DT, 0.200 DT, etc.) doit Ãªtre inclus dans Articles avec nom "TIMBRE FISCAL"
2. Si tu vois "100 DT", convertis-le en "0.100 DT" pour les timbres fiscaux
3. Retourne UNIQUEMENT le JSON, sans commentaires ni texte supplÃ©mentaire
4. Commence par {{ et termine par }}
"""
    try:
        print("Tentative avec modÃ¨le rapide (mistral)...")
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={"model": "mistral", "prompt": prompt, "stream": False},
            timeout=30
        )
        print("RÃ©ponse reÃ§ue de l'API Ollama (modÃ¨le rapide)")
        result_text = response.json().get("response", "")
        
        # Parser le JSON de la rÃ©ponse LLM
        parsed_data = clean_json_response(result_text)
        if parsed_data:
            result_data = {
                "Date": parsed_data.get("Date", ""),
                "Magasin": parsed_data.get("Magasin", ""),
                "NumeroTicket": parsed_data.get("NumeroTicket", ""),
                "Total": parsed_data.get("Total", ""),
                "Articles": parsed_data.get("Articles", []),
                "Commentaire": "DonnÃ©es extraites par modÃ¨le rapide",
                "texte_fusionne": result_text.strip()
            }
            
            # Post-traitement pour s'assurer que le timbre fiscal est bien dÃ©tectÃ©
            result_data = post_process_timbre_fiscal(result_data)
            
            # Validation et correction avec regex
            texte_ocr_combined = f"{docling_text}\n{tesseract_text}\n{doctr_text}"
            result_data = valider_et_corriger_avec_regex(result_data, texte_ocr_combined)
            
            return result_data
        
        # Si pas de JSON, retourner le texte brut
        return {
            "Date": "",
            "Magasin": "",
            "NumeroTicket": "",
            "Total": "",
            "Articles": [],
            "Commentaire": "Texte fusionnÃ© et corrigÃ© (modÃ¨le rapide)",
            "texte_fusionne": result_text.strip()
        }
    except requests.exceptions.Timeout:
        print("Timeout avec mistral, essai avec modÃ¨le ultra-rapide...")
        return analyze_three_texts_with_llm_ultra_fast(ocr_results)
    except Exception as e:
        print("Erreur avec modÃ¨le rapide:", str(e))
        return {"error": f"Erreur API (modÃ¨le rapide) : {str(e)}"}

def analyze_three_texts_with_llm_ultra_fast(ocr_results):
    """
    Version ultra-rapide avec un modÃ¨le trÃ¨s lÃ©ger
    """
    docling_text = ocr_results.get("docling", "")
    tesseract_text = ocr_results.get("tesseract", "")
    doctr_text = ocr_results.get("doctr", "")
    
    prompt = f"""Tu es un assistant expert en analyse de tickets de caisse.

Voici trois extraits OCR du mÃªme ticket :

--- OCR DocLing ---
{docling_text}

--- OCR Tesseract ---
{tesseract_text}

--- OCR Doctr ---
{doctr_text}

Extrais les Ã©lÃ©ments suivants et retourne UNIQUEMENT un objet JSON valide :

{{
  "Magasin": "Nom du magasin",
  "NumeroTicket": "NumÃ©ro du ticket",
  "Date": "JJ/MM/AAAA HH:MM",
  "Articles": [
    {{ "nom": "Nom article", "prix": "Prix en DT" }},
    {{ "nom": "TIMBRE FISCAL", "prix": "0.100 DT" }}
  ],
  "Total": "Montant total en DT"
}}

RÃˆGLES IMPORTANTES :
1. Le timbre fiscal (0.100 DT, 0.200 DT, etc.) doit Ãªtre inclus dans Articles avec nom "TIMBRE FISCAL"
2. Si tu vois "100 DT", convertis-le en "0.100 DT" pour les timbres fiscaux
3. Retourne UNIQUEMENT le JSON, sans commentaires ni texte supplÃ©mentaire
4. Commence par {{ et termine par }}
"""
    try:
        print("Tentative avec modÃ¨le ultra-rapide (llama2)...")
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={"model": "llama2", "prompt": prompt, "stream": False},
            timeout=15
        )
        print("RÃ©ponse reÃ§ue de l'API Ollama (modÃ¨le ultra-rapide)")
        result_text = response.json().get("response", "")
        
        # Parser le JSON de la rÃ©ponse LLM
        parsed_data = clean_json_response(result_text)
        if parsed_data:
            result_data = {
                "Date": parsed_data.get("Date", ""),
                "Magasin": parsed_data.get("Magasin", ""),
                "NumeroTicket": parsed_data.get("NumeroTicket", ""),
                "Total": parsed_data.get("Total", ""),
                "Articles": parsed_data.get("Articles", []),
                "Commentaire": "DonnÃ©es extraites par modÃ¨le ultra-rapide",
                "texte_fusionne": result_text.strip()
            }
            
            # Post-traitement pour s'assurer que le timbre fiscal est bien dÃ©tectÃ©
            result_data = post_process_timbre_fiscal(result_data)
            
            # Validation et correction avec regex
            texte_ocr_combined = f"{docling_text}\n{tesseract_text}\n{doctr_text}"
            result_data = valider_et_corriger_avec_regex(result_data, texte_ocr_combined)
            
            return result_data
        
        # Si pas de JSON, retourner le texte brut
        return {
            "Date": "",
            "Magasin": "",
            "NumeroTicket": "",
            "Total": "",
            "Articles": [],
            "Commentaire": "Texte fusionnÃ© et corrigÃ© (modÃ¨le ultra-rapide)",
            "texte_fusionne": result_text.strip()
        }
    except Exception as e:
        print("Erreur avec modÃ¨le ultra-rapide:", str(e))
        return {"error": f"Erreur API (modÃ¨le ultra-rapide) : {str(e)}"}

import os
import json
import re
from openai import OpenAI
try:
    import google.generativeai as genai
except Exception:
    genai = None
    print("Warning: google.generativeai package not installed — Gemini integration disabled.")
from dotenv import load_dotenv
load_dotenv()

# Initialiser le client OpenAI avec l'URL HuggingFace
try:
    hf_token = os.environ.get("HF_TOKEN", "")
    if not hf_token:
        print("âš ï¸ Token HuggingFace manquant dans .env")
        client = None
    else:
        client = OpenAI(
            base_url="https://router.huggingface.co/v1",
            api_key=hf_token,
        )
        print("âœ… Client HuggingFace initialisÃ©")
except Exception as e:
    print(f"âŒ Erreur initialisation client HuggingFace: {e}")
    client = None

def analyze_three_texts_with_llm(ocr_results):
    docling_text = ocr_results.get("docling", "")
    tesseract_text = ocr_results.get("tesseract", "")
    doctr_text = ocr_results.get("doctr", "")

    if not docling_text.strip() and not tesseract_text.strip() and not doctr_text.strip():
        return {
            "Date": "",
            "Magasin": "",
            "NumeroTicket": "",
            "Total": "",
            "Articles": [],
            "Commentaire": "Aucun texte OCR extrait - les textes sont vides"
        }

    prompt = f"""Tu es un assistant expert en analyse de tickets de caisse.

Voici un extrait OCR du ticket de caisse :

--- OCR Doctr ---
{doctr_text}

Ta tÃ¢che est d'extraire les Ã©lÃ©ments suivants et de retourner UNIQUEMENT un objet JSON valide :

{{
  "Magasin": "Nom du magasin",
  "NumeroTicket": "NumÃ©ro du ticket",
  "Date": "JJ/MM/AAAA HH:MM",
  "Articles": [
    {{ "nom": "Nom article", "prix": "Prix en DT" }},
    {{ "nom": "TIMBRE FISCAL", "prix": "0.100 DT" }}
  ],
  "Total": "Montant total en DT"
}}

RÃˆGLES IMPORTANTES :
1. Le timbre fiscal (0.100 DT, 0.200 DT, etc.) doit Ãªtre inclus dans Articles avec nom "TIMBRE FISCAL"
2. Si tu vois "100 DT", convertis-le en "0.100 DT" pour les timbres fiscaux
3. Retourne UNIQUEMENT le JSON, sans commentaires ni texte supplÃ©mentaire
4. Commence par {{ et termine par }}
5. Assure-toi que tous les montants soient des chaÃ®nes terminÃ©es par 'DT'

NE RENVOIE QUE UN OBJET JSON. PAS DE COMMENTAIRES, PAS DE TEXTE, PAS DE PENSÃ‰ES.
Commence DIRECTEMENT par '{' et termine par '}'.
"""

    try:
        # VÃ©rifier si le client est disponible
        if client is None:
            print("âŒ Client HuggingFace non disponible, utilisation du fallback Ollama")
            raise Exception("Client HuggingFace non initialisÃ©")
            
        print("ðŸ”„ Appel Ã  l'API HuggingFace avec Qwen...")

        completion = client.chat.completions.create(
            model="Qwen/Qwen3-30B-A3B:novita",
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0,
            timeout=30  # Timeout de 30 secondes
        )

        result_text = completion.choices[0].message.content.strip()

        # Clean and extract JSON with error handling
        try:
            # Utiliser la fonction de nettoyage amÃ©liorÃ©e
            parsed_data = clean_json_response(result_text)
            
            if not parsed_data:
                print("âŒ Aucun JSON valide trouvÃ© dans la rÃ©ponse LLM")
                return {
                    "error": "Invalid JSON response from LLM: No valid JSON found",
                    "raw_response": result_text,
                    "Commentaire": "La rÃ©ponse LLM ne contient pas de JSON valide"
                }
                
            print(f"âœ… JSON parsÃ© avec succÃ¨s")
            
            result_data = {
                "Date": parsed_data.get("Date", ""),
                "Magasin": parsed_data.get("Magasin", ""),
                "NumeroTicket": parsed_data.get("NumeroTicket", ""),
                "Total": parsed_data.get("Total", ""),
                "Articles": parsed_data.get("Articles", []),
                "Commentaire": "DonnÃ©es extraites via HuggingFace Qwen",
                "texte_fusionne": result_text
            }

            # Post-traitement personnalisÃ©
            result_data = post_process_timbre_fiscal(result_data)
            texte_ocr_combined = f"{docling_text}\n{tesseract_text}\n{doctr_text}"
            result_data = valider_et_corriger_avec_regex(result_data, texte_ocr_combined)

            return result_data
            
        except Exception as e:
            logger.error(f"Erreur lors du parsing JSON: {str(e)}")
            return {
                "error": f"Invalid JSON response from LLM: {str(e)}",
                "raw_response": result_text,
                "Commentaire": f"Erreur lors du parsing JSON: {str(e)}"
            }

    except Exception as e:
        print("Erreur lors de l'appel Ã  l'API HuggingFace:", str(e))
        # Fallback vers les modÃ¨les Ollama locaux
        try:
            print("Tentative de fallback vers Ollama...")
            return analyze_three_texts_with_llm_fast(ocr_results)
        except Exception as fallback_error:
            print(f"Erreur avec fallback Ollama: {str(fallback_error)}")
            # Dernier recours : analyse simple avec regex
            try:
                texte_ocr_combined = f"{docling_text}\n{tesseract_text}\n{doctr_text}"
                regex_result = extraire_elements_avec_regex(texte_ocr_combined)
                
                return {
                    "Date": regex_result.get("dates_valides", [""])[0] if regex_result.get("dates_valides") else "",
                    "Magasin": "",
                    "NumeroTicket": "",
                    "Total": regex_result.get("total", ""),
                    "Articles": regex_result.get("articles", []),
                    "Commentaire": f"Analyse par regex (fallback) - Erreur API: {str(e)}",
                    "texte_fusionne": texte_ocr_combined,
                    "ValidationRegex": {
                        "total_coherent": regex_result.get("total_coherent", False),
                        "somme_articles": regex_result.get("somme_articles", ""),
                        "total_detecte": regex_result.get("total", "")
                    }
                }
            except Exception as regex_error:
                return {
                    "Date": "",
                    "Magasin": "",
                    "NumeroTicket": "",
                    "Total": "",
                    "Articles": [],
                    "Commentaire": f"Erreur complÃ¨te - API: {str(e)}, Fallback: {str(fallback_error)}, Regex: {str(regex_error)}",
                    "texte_fusionne": "Erreur de traitement"
                }

def analyze_three_texts_with_gemini(ocr_results):
    """
    Analyse les 3 textes OCR avec Google Generative AI (Gemini)
    """
    try:
        # Configuration de l'API Google Generative AI
        api_key = os.getenv('GOOGLE_API_KEY')
        if not api_key or api_key == 'your_google_api_key_here':
            # Mode dÃ©mo - retourner des donnÃ©es d'exemple
            logger.warning("ClÃ© API Google Generative AI manquante - Mode dÃ©mo activÃ©")
            return {
                "error": "ClÃ© API Google Generative AI manquante. Voici un exemple de rÃ©sultat.",
                "demo_mode": True,
                "Magasin": "DEMO - Magasin Exemple",
                "Date": "2024-01-15",
                "Heure": "14:30",
                "Numero_ticket": "DEMO-001",
                "Articles": [
                    {
                        "nom": "Article dÃ©mo 1",
                        "quantite": 2,
                        "prix_unitaire": 5.500,
                        "prix_total": 11.000
                    },
                    {
                        "nom": "TIMBRE FISCAL",
                        "quantite": 1,
                        "prix_unitaire": 0.100,
                        "prix_total": 0.100
                    }
                ],
                "Sous_total": 11.000,
                "Remise": 0.000,
                "Timbre_fiscal": 0.100,
                "Total": 11.100,
                "Methode_paiement": "espÃ¨ces",
                "TVA_details": {
                    "taux_19": 2.090,
                    "taux_13": 0.000,
                    "taux_7": 0.000
                },
                "model_used": "Google Gemini 1.5 Flash (Mode DÃ©mo)",
                "analysis_timestamp": datetime.now().isoformat(),
                "raw_response": "Mode dÃ©mo - Aucune API appelÃ©e",
                "instructions": {
                    "title": "Comment obtenir une clÃ© API Google Generative AI :",
                    "steps": [
                        "1. Allez sur https://makersuite.google.com/app/apikey",
                        "2. Connectez-vous avec votre compte Google",
                        "3. Cliquez sur 'Create API Key'",
                        "4. Copiez la clÃ© gÃ©nÃ©rÃ©e",
                        "5. Remplacez 'your_google_api_key_here' dans le fichier .env",
                        "6. RedÃ©marrez le serveur Django"
                    ]
                }
            }
        
        if genai is None:
            logger.warning("Gemini requested but google.generativeai client is not available; returning demo response")
            # fall back to demo response handled by existing logic above
            return {
                "error": "Clé API Google Generative AI manquante ou client indisponible. Mode démo activé.",
                "demo_mode": True,
            }

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        # Fusionner les 3 textes OCR
        texte_fusionne = f"""
        === TEXTE DOCTR ===
        {ocr_results.get('doctr', '')}
        
        === TEXTE TESSERACT ===
        {ocr_results.get('tesseract', '')}
        
        === TEXTE DOCLING ===
        {ocr_results.get('docling', '')}
        """
        
        # Prompt simplifiÃ© pour correspondre au format Qwen
        prompt = f"""
        Tu es un expert en extraction de donnÃ©es de tickets de caisse. Analyse ces 3 textes OCR d'un mÃªme ticket et extrais les informations suivantes au format JSON strict :
        
        {{
            "Magasin": "nom du magasin",
            "Date": "YYYY-MM-DD",
            "NumeroTicket": "numÃ©ro du ticket" ou null,
            "Articles": [
                {{
                    "nom": "nom de l'article",
                    "prix": "prix de l'article avec unitÃ© (ex: 5.500 DT)"
                }}
            ],
            "Total": "montant total avec unitÃ© (ex: 39.500 DT)"
        }}
        
        RÃˆGLES IMPORTANTES :
        - Utilise les 3 textes pour obtenir la meilleure prÃ©cision
        - Corrige les erreurs OCR courantes : Oâ†’0, Iâ†’1, etc.
        - Si une information n'est pas trouvÃ©e, utilise null
        - RÃ©ponds UNIQUEMENT avec le JSON, sans texte supplÃ©mentaire
        
        TEXTES OCR Ã€ ANALYSER :
        {texte_fusionne}
        """
        
        logger.info("Envoi de la requÃªte Ã  Google Generative AI...")
        
        # GÃ©nÃ©rer la rÃ©ponse
        response = model.generate_content(prompt)
        raw_response = response.text
        
        logger.info(f"RÃ©ponse brute reÃ§ue de Gemini: {raw_response[:200]}...")
        
        # Nettoyer et parser la rÃ©ponse JSON
        result_data = clean_json_response(raw_response)
        
        if result_data and isinstance(result_data, dict):
            logger.info("Analyse Gemini rÃ©ussie")
            
            # Ajouter des mÃ©tadonnÃ©es
            result_data['texte_fusionne'] = texte_fusionne
            result_data['raw_response'] = raw_response
            result_data['model_used'] = 'Google Gemini 1.5 Flash'
            result_data['analysis_timestamp'] = datetime.now().isoformat()
            
            # Validation et correction avec regex
            result_data = valider_et_corriger_avec_regex(result_data, texte_fusionne)
            
            return result_data
        else:
            logger.error("Erreur de parsing JSON Gemini: RÃ©ponse invalide")
            return {
                "error": "RÃ©ponse JSON invalide de Gemini",
                "raw_response": raw_response,
                "parsed_response": result_data
            }
    
    except Exception as e:
        logger.error(f"Erreur lors de l'analyse avec Gemini: {str(e)}")
        return {
            "error": f"Erreur lors de l'analyse Gemini: {str(e)}",
            "raw_response": None
        }

def extraire_elements_avec_regex(texte):
    """
    Extraction prÃ©cise d'Ã©lÃ©ments avec regex pour validation/correction des rÃ©sultats LLM
    """
    if not texte or not isinstance(texte, str):
        return {}
    
    # Regex pour dates (JJ/MM/AAAA ou JJ-MM-AAAA)
    dates = re.findall(r'\b([0-3]?\d)[/-]([0-1]?\d)[/-](\d{4})\b', texte)
    dates_valides = []
    for d, m, y in dates:
        try:
            date_obj = datetime.strptime(f"{d}/{m}/{y}", "%d/%m/%Y")
            dates_valides.append(date_obj.strftime("%d/%m/%Y"))
        except:
            pass
    
    # Regex pour prix au format comme 0.100 DT, 12.50 DT
    prix_pattern = r'\b\d+\.\d{2,3}\s?DT\b'
    prix_trouves = re.findall(prix_pattern, texte)
    
    # Convertir les prix en float pour calcul (supprimer " DT" et convertir)
    def prix_to_float(p):
        try:
            return float(p.replace("DT", "").strip())
        except Exception:
            return 0.0
    
    prix_floats = [prix_to_float(p) for p in prix_trouves]

    # Regex pour dÃ©tecter timbre fiscal (ex: 0.100 DT, 0.200 DT etc.)
    timbres = [p for p in prix_trouves if p in ["0.100 DT", "0.200 DT", "0.300 DT", "0.400 DT", "0.500 DT"]]
    timbres_float = [prix_to_float(t) for t in timbres]
    
    # Supposons que le total est indiquÃ© par une ligne avec "Total" + montant
    total_match = re.search(r'(Total|TOTAL|total)\s*[:\-]?\s*(\d+\.\d{2,3})\s?DT', texte)
    total = None
    if total_match:
        try:
            total = float(total_match.group(2))
        except Exception:
            total = None
    
    # Extraire les articles, en filtrant les mots interdits (espÃ¨ce, rendu, reÃ§u, total, etc.)
    # On suppose que chaque ligne contient nom article + prix (ex: Pain 1.200 DT)
    mots_interdits = ['espece', 'rendu', 'recu', 'total', 'remise', 'timbre', 'fiscal', 'taxe', 'stamp','pece']
    
    articles = []
    # Exemple simple : chercher toutes les lignes avec un prix, puis extraire le nom avant
    lignes = texte.split('\n')
    for ligne in lignes:
        prix_match = re.search(prix_pattern, ligne)
        if prix_match:
            prix = prix_match.group()
            # Extraire nom avant prix
            nom = ligne[:prix_match.start()].strip().lower()
            
            # VÃ©rifier si un mot interdit est dans le nom
            if any(mot in nom for mot in mots_interdits):
                continue  # Ignore cette ligne, ce n'est pas un article
            
            # On suppose que le nom est le texte avant le prix
            nom_article = ligne[:prix_match.start()].strip()
            articles.append({"nom": nom_article, "prix": prix})

    # VÃ©rification heuristique : somme des articles == total
    somme_articles = sum([prix_to_float(a["prix"]) for a in articles])
    total_correct = (total is not None and abs(somme_articles - total) < 0.01)  # marge d'erreur 1 centime
    
    resultat = {
        "dates_valides": dates_valides,
        "timbres_fiscaux": timbres,
        "total": f"{total:.3f} DT" if total is not None else "",
        "articles": articles,
        "somme_articles": f"{somme_articles:.3f} DT",
        "total_coherent": total_correct
    }
    
    if not total_correct:
        # Previously we exposed a user-facing alert when the sum of detected articles
        # didn't match the detected total. That banner has been removed from the UI
        # per request; keep the info only in logs for debugging if needed.
        total_str = f"{total:.3f}" if total is not None else "0.000"
        # logger.debug could be used; use print for parity with existing debug style
        print(f"Validation mismatch: somme_articles={somme_articles:.3f} DT, total={total_str} DT")
    
    return resultat

def post_process_timbre_fiscal(result_data):
    """
    Post-traitement pour s'assurer que le timbre fiscal est bien dÃ©tectÃ© et inclus dans les articles
    """
    if not result_data or not isinstance(result_data, dict):
        return result_data
    
    articles = result_data.get("Articles", [])
    
    # Chercher le timbre fiscal dans les articles
    for article in articles:
        if isinstance(article, dict):
            nom = article.get("nom", "").lower()
            prix = article.get("prix", "")
            
            # DÃ©tecter le timbre fiscal par le montant (0.100 DT, 0.200 DT, etc.)
            if prix in ["0.100 DT", "0.200 DT", "0.300 DT", "0.400 DT", "0.500 DT"]:
                # S'assurer que le nom est "TIMBRE FISCAL"
                if "timbre" not in nom and "fiscal" not in nom:
                    article["nom"] = "TIMBRE FISCAL"
                    print(f"Nom de l'article timbre fiscal corrigÃ©: {article['nom']}")
                break
            
            # DÃ©tecter par le nom aussi
            if any(keyword in nom for keyword in ["timbre", "fiscal", "taxe", "stamp"]):
                # S'assurer que le nom est "TIMBRE FISCAL"
                article["nom"] = "TIMBRE FISCAL"
                print(f"Nom de l'article timbre fiscal corrigÃ©: {article['nom']}")
                break
            
            # Conversion des montants : 100 DT â†’ 0.100 DT pour les timbres fiscaux
            if prix == "100 DT" and any(keyword in nom for keyword in ["timbre", "fiscal", "taxe", "stamp"]):
                article["prix"] = "0.100 DT"
                article["nom"] = "TIMBRE FISCAL"
                print(f"Montant timbre fiscal converti: 100 DT â†’ 0.100 DT")
                break
    
    return result_data

def valider_et_corriger_avec_regex(result_data, texte_ocr):
    """
    Valide et corrige les rÃ©sultats LLM avec l'extraction regex
    """
    if not result_data or not isinstance(result_data, dict):
        return result_data
    
    # Extraire les Ã©lÃ©ments avec regex
    regex_result = extraire_elements_avec_regex(texte_ocr)
    
    print("=== VALIDATION REGEX ===")
    print(f"Dates trouvÃ©es: {regex_result.get('dates_valides', [])}")
    print(f"Timbres fiscaux: {regex_result.get('timbres_fiscaux', [])}")
    print(f"Total regex: {regex_result.get('total', '')}")
    print(f"Articles regex: {len(regex_result.get('articles', []))}")
    print(f"Total cohÃ©rent: {regex_result.get('total_coherent', False)}")
    print("========================")
    
    # Corriger la date si nÃ©cessaire
    if not result_data.get("Date") and regex_result.get("dates_valides"):
        result_data["Date"] = regex_result["dates_valides"][0]
        print(f"Date corrigÃ©e avec regex: {result_data['Date']}")
    
    
    
    # Corriger le total si nÃ©cessaire
    if not result_data.get("Total") and regex_result.get("total"):
        result_data["Total"] = regex_result["total"]
        print(f"Total corrigÃ© avec regex: {result_data['Total']}")
    
    # Corriger les articles si nÃ©cessaire (si le LLM n'a pas bien dÃ©tectÃ©)
    if not result_data.get("Articles") and regex_result.get("articles"):
        result_data["Articles"] = regex_result["articles"]
        print(f"Articles corrigÃ©s avec regex: {len(result_data['Articles'])} articles")
    
    # Ajouter des informations de validation
    if regex_result.get("alerte"):
        result_data["AlerteValidation"] = regex_result["alerte"]
        print(f"Alerte ajoutÃ©e: {regex_result['alerte']}")
    
    result_data["ValidationRegex"] = {
        "total_coherent": regex_result.get("total_coherent", False),
        "somme_articles": regex_result.get("somme_articles", ""),
        "total_detecte": regex_result.get("total", "")
    }
    
    return result_data

def extract_text_doctr(file_path):
    try:
        print(f"Doctr: Processing {file_path}")

        # Vérifier que le fichier existe
        if not file_path or not os.path.exists(file_path):
            error_msg = f"Erreur Doctr: fichier introuvable: {file_path}"
            print(error_msg)
            return error_msg

        ext = os.path.splitext(file_path)[1].lower()

        images = []
        # Si c'est un PDF, convertir en images (pages)
        if ext == '.pdf':
            try:
                from pdf2image import convert_from_path
            except Exception as e:
                error_msg = f"Erreur Doctr: pdf2image non disponible pour convertir le PDF: {e}"
                print(error_msg)
                return error_msg

            try:
                images = convert_from_path(file_path)
                if not images:
                    raise ValueError('Aucune page convertie depuis le PDF')
            except Exception as e:
                error_msg = f"Erreur Doctr: impossible de convertir le PDF en images: {e}"
                print(error_msg)
                return error_msg

        else:
            # Traiter comme image (JPG, PNG, etc.)
            try:
                from PIL import Image
            except Exception as e:
                error_msg = f"Erreur Doctr: PIL non disponible pour ouvrir l'image: {e}"
                print(error_msg)
                return error_msg

            try:
                img = Image.open(file_path).convert('RGB')
                images = [img]
            except Exception as e:
                error_msg = f"Erreur Doctr: impossible d'ouvrir l'image: {e}"
                print(error_msg)
                return error_msg

        # Construire un DocumentFile à partir des images PIL
        try:
            doc = DocumentFile.from_images(images)
        except Exception as e:
            # Fallback: essayer de passer les chemins si la conversion échoue
            try:
                doc = DocumentFile.from_images([file_path])
            except Exception as e2:
                error_msg = f"Erreur Doctr: unable to build DocumentFile: {e}; fallback error: {e2}"
                print(error_msg)
                return error_msg

        # Charger le modèle Doctr
        model = ocr_predictor(pretrained=True)

        # Faire la prédiction
        result = model(doc)

        # Extraire le texte brut
        extracted_text = result.render()
        print(f"Doctr: Extracted {len(extracted_text)} characters")
        return extracted_text

    except Exception as e:
        error_msg = f"Erreur Doctr inattendue: {str(e)}"
        print(error_msg)
        return error_msg

def extract_text_docling(file_path):
    try:
        print(f"Docling: Processing {file_path}")
        converter = DocumentConverter()
        path_obj = Path(file_path)
        result = converter.convert(path_obj)
        document = result.document

        text_lines = {}
        if hasattr(document, 'texts'):
            for text_item in document.texts:
                if hasattr(text_item, 'prov') and text_item.prov:
                    y_pos = sum([prov.bbox.t for prov in text_item.prov]) / len(text_item.prov)
                    text_lines[y_pos] = text_lines.get(y_pos, "") + " " + text_item.text

        sorted_lines = sorted(text_lines.items(), key=lambda x: -x[0])
        extracted_text = "\n".join([line[1].strip() for line in sorted_lines])
        print(f"Docling: Extracted {len(extracted_text)} characters")
        return extracted_text

    except Exception as e:
        error_msg = f"Erreur Docling: {str(e)}"
        print(error_msg)
        return error_msg

def extract_text_tesseract(file_path):
    try:
        print(f"Tesseract: Processing {file_path}")
        
        # VÃ©rifier si Tesseract est disponible
        try:
            import pytesseract
            # Test rapide pour vÃ©rifier si Tesseract est installÃ©
            pytesseract.get_tesseract_version()
        except Exception as tesseract_error:
            error_msg = f"Tesseract non disponible: {str(tesseract_error)}\n\nPour installer Tesseract:\n1. TÃ©lÃ©chargez: https://github.com/UB-Mannheim/tesseract/wiki\n2. Installez dans C:\\Program Files\\Tesseract-OCR\\\n3. Ajoutez au PATH systÃ¨me"
            print(error_msg)
            return error_msg
        
        ext = os.path.splitext(file_path)[1].lower()
        images = []
        if ext == '.pdf':
            try:
                from pdf2image import convert_from_path
                images = convert_from_path(file_path)
            except ImportError:
                return "Erreur: pdf2image non installÃ© pour traiter les PDF"
        else:
            images = [Image.open(file_path)]
        
        text = ""
        for img in images:
            text += pytesseract.image_to_string(img, lang='fra+eng') + "\n"
        extracted_text = text.strip()
        print(f"Tesseract: Extracted {len(extracted_text)} characters")
        return extracted_text
    except Exception as e:
        error_msg = f"Erreur Tesseract: {str(e)}"
        print(error_msg)
        return error_msg

def save_ticket_to_history(llm_analysis):
    """
    Sauvegarde un ticket analysÃ© dans l'historique
    """
    if not llm_analysis or not isinstance(llm_analysis, dict):
        return None
    
    try:
        # Extraire les donnÃ©es du ticket
        date_str = llm_analysis.get("Date", "")
        magasin = llm_analysis.get("Magasin", "Magasin inconnu")
        total_str = llm_analysis.get("Total", "0.000 DT")
        numero_ticket = llm_analysis.get("NumeroTicket", "")
        articles = llm_analysis.get("Articles", [])
        
        # Nettoyer le total
        try:
            total_clean = total_str.replace(" DT", "").replace(",", ".")
            total_decimal = Decimal(total_clean)
        except:
            total_decimal = Decimal('0.000')
        
        # Parser la date
        try:
            if date_str:
                # Essayer diffÃ©rents formats de date
                date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]
                ticket_date = None
                for fmt in date_formats:
                    try:
                        ticket_date = datetime.strptime(date_str.split()[0], fmt).date()
                        break
                    except:
                        continue
                if not ticket_date:
                    ticket_date = date.today()
            else:
                ticket_date = date.today()
        except:
            ticket_date = date.today()
        
        # CrÃ©er l'entrÃ©e dans l'historique
        ticket_history = TicketHistory.objects.create(
            date_ticket=ticket_date,
            magasin=magasin,
            total=total_decimal,
            numero_ticket=numero_ticket,
            articles_data=articles,
            llm_analysis=llm_analysis
        )
        
        print(f"Ticket sauvegardÃ©: {ticket_history}")
        return ticket_history
        
    except Exception as e:
        print(f"Erreur lors de la sauvegarde du ticket: {e}")
        return None

def generate_accounting_report(llm_analysis, compte="606100", description="Achat divers", save_to_db=False):
    """
    GÃ©nÃ¨re un rapport comptable Ã  partir des donnÃ©es LLM
    """
    if not llm_analysis or not isinstance(llm_analysis, dict):
        return None
    
    # Extraire les donnÃ©es du ticket
    date_ticket = llm_analysis.get("Date", "")
    magasin = llm_analysis.get("Magasin", "Magasin inconnu")
    total = llm_analysis.get("Total", "0.000 DT")
    
    # Nettoyer le total (enlever "DT" et convertir en float)
    try:
        total_clean = total.replace(" DT", "").replace(",", ".")
        total_float = float(total_clean)
    except:
        total_float = 0.0
    
    # CrÃ©er le libellÃ© d'Ã©criture
    libelle_ecriture = f"Achat-{magasin}"
    
    # CrÃ©er le rapport comptable
    report_data = {
        "date_ticket": date_ticket,
        "compte": compte,
        "description": description,
        "libelle_ecriture": libelle_ecriture,
        "debit": f"{total_float:.3f}",
        "credit": "",
        "magasin": magasin,
        "total_original": total
    }
    
    # Sauvegarder dans la base de donnÃ©es si demandÃ©
    if save_to_db:
        try:
            # Sauvegarder le ticket dans l'historique
            ticket_history = save_ticket_to_history(llm_analysis)
            
            if ticket_history:
                # CrÃ©er l'entrÃ©e comptable
                try:
                    # Parser la date pour l'Ã©criture comptable
                    date_ecriture = date.today()
                    if date_ticket:
                        date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]
                        for fmt in date_formats:
                            try:
                                date_ecriture = datetime.strptime(date_ticket.split()[0], fmt).date()
                                break
                            except:
                                continue
                    
                    accounting_entry = AccountingEntry.objects.create(
                        ticket=ticket_history,
                        date_ecriture=date_ecriture,
                        compte=compte,
                        description=description,
                        libelle_ecriture=libelle_ecriture,
                        debit=Decimal(f"{total_float:.3f}"),
                        credit=None
                    )
                    print(f"EntrÃ©e comptable crÃ©Ã©e: {accounting_entry}")
                except Exception as e:
                    print(f"Erreur lors de la crÃ©ation de l'entrÃ©e comptable: {e}")
        except Exception as e:
            print(f"Erreur lors de la sauvegarde: {e}")
    
    return report_data

def generate_accounting_pdf(report_data):
    """
    GÃ©nÃ¨re un PDF du bilan comptable
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Titre
    title = Paragraph("Bilan Comptable - Ticket de Caisse", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Informations du ticket
    ticket_info = [
        ["Date du ticket:", report_data["date_ticket"]],
        ["Magasin:", report_data["magasin"]],
        ["Total:", report_data["total_original"]],
    ]
    
    ticket_table = Table(ticket_info, colWidths=[2*inch, 4*inch])
    ticket_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(ticket_table)
    elements.append(Spacer(1, 20))
    
    # Tableau comptable
    accounting_data = [
        ["Date", "Compte", "Description", "LibellÃ© Ã‰criture", "DÃ©bit", "CrÃ©dit"],
        [
            report_data["date_ticket"],
            report_data["compte"],
            report_data["description"],
            report_data["libelle_ecriture"],
            f"{report_data['debit']} DT",
            report_data["credit"]
        ]
    ]
    
    accounting_table = Table(accounting_data, colWidths=[1*inch, 1*inch, 1.5*inch, 2*inch, 1*inch, 1*inch])
    accounting_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, 1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(accounting_table)
    
    # Construire le PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_accounting_excel(report_data):
    """
    GÃ©nÃ¨re un fichier Excel du bilan comptable
    """
    # CrÃ©er un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan Comptable"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Titre principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "Bilan Comptable - Ticket de Caisse"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_alignment
    
    # Informations du ticket (ligne 3-5)
    ws['A3'] = "Date du ticket:"
    ws['B3'] = report_data["date_ticket"]
    ws['A4'] = "Magasin:"
    ws['B4'] = report_data["magasin"]
    ws['A5'] = "Total:"
    ws['B5'] = report_data["total_original"]
    
    # Style pour les labels
    for row in range(3, 6):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # En-tÃªtes du tableau comptable (ligne 7)
    headers = ["Date", "Compte", "Description", "LibellÃ© Ã‰criture", "DÃ©bit", "CrÃ©dit"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # DonnÃ©es comptables (ligne 8)
    data_row = [
        report_data["date_ticket"],
        report_data["compte"],
        report_data["description"],
        report_data["libelle_ecriture"],
        f"{report_data['debit']} DT",
        report_data["credit"]
    ]
    
    for col, value in enumerate(data_row, 1):
        cell = ws.cell(row=8, column=col, value=value)
        cell.border = border
        cell.alignment = center_alignment
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 12, 20, 25, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_cumulative_accounting_pdf(start_date=None, end_date=None):
    """
    GÃ©nÃ¨re un PDF cumulatif de toutes les entrÃ©es comptables
    """
    # RÃ©cupÃ©rer toutes les entrÃ©es comptables
    entries = AccountingEntry.objects.all()
    
    if start_date:
        entries = entries.filter(date_ecriture__gte=start_date)
    if end_date:
        entries = entries.filter(date_ecriture__lte=end_date)
    
    entries = entries.order_by('date_ecriture')

    # Exclure deux lignes spÃ©cifiques demandÃ©es par l'utilisateur
    try:
        from datetime import date as _date
        entries = entries.exclude(
            compte='531200',
            description='Paiement ticket de caisse',
            libelle_ecriture='Paiement ticket - AZIZA - 80102080 - 4.090 DT',
            date_ecriture__in=[_date(2025, 2, 1), _date(2025, 8, 8)]
        )
    except Exception:
        pass  # Ne jamais bloquer la gÃ©nÃ©ration si exclusion Ã©choue
    
    if not entries.exists():
        return None
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Titre
    title = Paragraph("Bilan Comptable Cumulatif", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Informations gÃ©nÃ©rales
    total_debit = sum(entry.debit for entry in entries)
    nb_entries = entries.count()
    
    summary_info = [
        ["PÃ©riode:", f"{start_date or 'DÃ©but'} Ã  {end_date or 'Aujourd\'hui'}"],
        ["Nombre d'Ã©critures:", str(nb_entries)],
        ["Total dÃ©bit:", f"{total_debit:.3f} DT"],
    ]
    
    summary_table = Table(summary_info, colWidths=[2*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Tableau des Ã©critures comptables
    accounting_data = [
        ["Date", "Compte", "Description", "LibellÃ© Ã‰criture", "DÃ©bit", "CrÃ©dit"]
    ]
    
    for entry in entries:
        accounting_data.append([
            entry.date_ecriture.strftime("%d/%m/%Y"),
            entry.compte,
            entry.description,
            entry.libelle_ecriture,
            f"{entry.debit:.3f} DT",
            f"{entry.credit:.3f} DT" if entry.credit else ""
        ])
    
    # Ajouter une ligne de total
    accounting_data.append([
        "", "", "", "TOTAL", f"{total_debit:.3f} DT", ""
    ])
    
    accounting_table = Table(accounting_data, colWidths=[1*inch, 1*inch, 1.5*inch, 2*inch, 1*inch, 1*inch])
    accounting_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(accounting_table)
    
    # Construire le PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_cumulative_accounting_excel(start_date=None, end_date=None):
    """
    GÃ©nÃ¨re un fichier Excel cumulatif de toutes les entrÃ©es comptables
    """
    # RÃ©cupÃ©rer toutes les entrÃ©es comptables
    entries = AccountingEntry.objects.all()
    
    if start_date:
        entries = entries.filter(date_ecriture__gte=start_date)
    if end_date:
        entries = entries.filter(date_ecriture__lte=end_date)
    
    entries = entries.order_by('date_ecriture')

    # Exclure deux lignes spÃ©cifiques demandÃ©es par l'utilisateur
    try:
        from datetime import date as _date
        entries = entries.exclude(
            compte='531200',
            description='Paiement ticket de caisse',
            libelle_ecriture='Paiement ticket - AZIZA - 80102080 - 4.090 DT',
            date_ecriture__in=[_date(2025, 2, 1), _date(2025, 8, 8)]
        )
    except Exception:
        pass
    
    if not entries.exists():
        return None
    
    # CrÃ©er un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan Comptable Cumulatif"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    total_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Titre principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "Bilan Comptable Cumulatif"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_alignment
    
    # Informations gÃ©nÃ©rales
    total_debit = sum(entry.debit for entry in entries)
    nb_entries = entries.count()
    
    ws['A3'] = "PÃ©riode:"
    ws['B3'] = f"{start_date or 'DÃ©but'} Ã  {end_date or 'Aujourd\'hui'}"
    ws['A4'] = "Nombre d'Ã©critures:"
    ws['B4'] = str(nb_entries)
    ws['A5'] = "Total dÃ©bit:"
    ws['B5'] = f"{total_debit:.3f} DT"
    
    # Style pour les labels
    for row in range(3, 6):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # En-tÃªtes du tableau comptable (ligne 7)
    headers = ["Date", "Compte", "Description", "LibellÃ© Ã‰criture", "DÃ©bit", "CrÃ©dit"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # DonnÃ©es comptables
    current_row = 8
    for entry in entries:
        data_row = [
            entry.date_ecriture.strftime("%d/%m/%Y"),
            entry.compte,
            entry.description,
            entry.libelle_ecriture,
            f"{entry.debit:.3f} DT",
            f"{entry.credit:.3f} DT" if entry.credit else ""
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            cell.alignment = center_alignment
        
        current_row += 1
    
    # Ligne de total
    total_row = ["", "", "", "TOTAL", f"{total_debit:.3f} DT", ""]
    for col, value in enumerate(total_row, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.border = border
        cell.alignment = center_alignment
        cell.fill = total_fill
        cell.font = Font(bold=True)
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 12, 20, 25, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def download_cumulative_excel(request):
    """
    Vue pour tÃ©lÃ©charger le bilan comptable cumulatif en Excel
    """
    if request.method == 'POST':
        try:
            # RÃ©cupÃ©rer les dates de filtrage
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            
            start_date = None
            end_date = None
            
            if start_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            if end_date_str:
                try:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            # GÃ©nÃ©rer le fichier Excel cumulatif
            excel_buffer = generate_cumulative_accounting_excel(start_date, end_date)
            
            if not excel_buffer:
                return HttpResponse("Aucune donnÃ©e disponible pour la pÃ©riode sÃ©lectionnÃ©e", status=400)
            
            # CrÃ©er la rÃ©ponse HTTP
            response = HttpResponse(
                excel_buffer.getvalue(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="bilan_comptable_cumulatif_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
            
        except Exception as e:
            print(f"Erreur lors de la gÃ©nÃ©ration du fichier Excel cumulatif: {e}")
            return HttpResponse(f"Erreur lors de la gÃ©nÃ©ration du fichier Excel: {str(e)}", status=500)
    
    return HttpResponse("MÃ©thode non autorisÃ©e", status=405)

def view_history(request):
    """
    Vue pour afficher l'historique des tickets avec gestion de budget
    """
    from datetime import datetime
    from .models import Budget
    
    # Précharger les écritures comptables pour éviter les requêtes N+1
    tickets = TicketHistory.objects.all().prefetch_related('accounting_entries').order_by('-created_at')
    total_amount = sum(ticket.total for ticket in tickets)
    
    # Gestion du budget
    budget_type = request.GET.get('budget_type', 'monthly')
    current_date = datetime.now()
    
    # RÃ©cupÃ©rer les paramÃ¨tres de mois et annÃ©e
    selected_year = int(request.GET.get('year', current_date.year))
    selected_month = int(request.GET.get('month', current_date.month))

    # Plage dynamique d'années (à partir de 2010 jusqu'à année courante + 5)
    start_year_range = 2010
    end_year_range = current_date.year + 5
    year_range = list(range(start_year_range, end_year_range + 1))
    
    # Calculer les dÃ©penses selon le type de budget
    if budget_type == 'monthly':
        # DÃ©penses du mois sÃ©lectionnÃ©
        monthly_tickets = tickets.filter(
            date_ticket__year=selected_year,
            date_ticket__month=selected_month
        )
        current_expenses = sum(ticket.total for ticket in monthly_tickets)
        # RÃ©cupÃ©rer le budget pour le mois/annÃ©e spÃ©cifique
        try:
            budget = Budget.objects.get(
                type_budget='monthly',
                annee=selected_year,
                mois=selected_month
            )
        except Budget.DoesNotExist:
            budget = None
    else:  # yearly
        # DÃ©penses de l'annÃ©e sÃ©lectionnÃ©e
        yearly_tickets = tickets.filter(date_ticket__year=selected_year)
        current_expenses = sum(ticket.total for ticket in yearly_tickets)
        # RÃ©cupÃ©rer le budget pour l'annÃ©e spÃ©cifique
        try:
            budget = Budget.objects.get(
                type_budget='yearly',
                annee=selected_year
            )
        except Budget.DoesNotExist:
            budget = None
    
    # Calculer les indicateurs de budget
    budget_info = {
        'budget_type': budget_type,
        'budget_amount': budget.montant if budget else None,
        'current_expenses': current_expenses,
        'remaining': None,
        'exceeded': None,
        'budget_exists': budget is not None,
        'current_year': selected_year,
        'current_month': selected_month,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'year_range': year_range,
        'start_year_range': start_year_range,
        'end_year_range': end_year_range
    }
    
    if budget:
        if current_expenses <= budget.montant:
            budget_info['remaining'] = budget.montant - current_expenses
        else:
            budget_info['exceeded'] = current_expenses - budget.montant
    
    # Identifier les tickets qui causent un dÃ©passement - VERSION OPTIMISÃ‰E
    tickets_with_budget_status = []
    running_total = 0
    budget_exceeded = False
    overflow_threshold_reached = False
    
    if budget:
        # Filtrer et trier les tickets selon la pÃ©riode sÃ©lectionnÃ©e
        if budget_type == 'monthly':
            period_tickets = tickets.filter(
                date_ticket__year=selected_year,
                date_ticket__month=selected_month
            )
        else:  # yearly
            period_tickets = tickets.filter(
                date_ticket__year=selected_year
            )
        
        # Trier par date et heure de crÃ©ation pour un ordre chronologique prÃ©cis
        sorted_tickets = period_tickets.order_by('date_ticket', 'created_at')
        
        for ticket in sorted_tickets:
            previous_total = running_total
            running_total += ticket.total
            
            # DÃ©terminer le statut du ticket
            ticket_status = {
                'ticket': ticket,
                'running_total': running_total,
                'causes_overflow': False,
                'is_over_budget': False,
                'percentage_of_budget': (running_total / budget.montant) * 100,
                'compte': None,
                'description': None,
            }

            # Priorité: utiliser les valeurs persistées dans llm_analysis si disponibles (compte/description)
            try:
                llm = ticket.llm_analysis or {}
                if isinstance(llm, dict) and (llm.get('compte') or llm.get('description')):
                    if llm.get('compte'):
                        ticket_status['compte'] = llm.get('compte')
                    if llm.get('description'):
                        ticket_status['description'] = llm.get('description')
                else:
                    # Récupérer un numéro de compte représentatif (priorité: première écriture débit > 0, sinon première écriture)
                    entries_related = ticket.accounting_entries.all()
                    compte_val = None
                    for e in entries_related:
                        if e.debit and e.debit > 0:
                            compte_val = e.compte
                            break
                    if not compte_val and entries_related:
                        compte_val = entries_related[0].compte
                        # If description not set from llm, use first related entry description
                        if not ticket_status.get('description') and entries_related:
                            ticket_status['description'] = entries_related[0].description
                    ticket_status['compte'] = compte_val
            except Exception:
                ticket_status['compte'] = None
            
            # Marquer si ce ticket spÃ©cifique cause le dÃ©passement
            if previous_total <= budget.montant and running_total > budget.montant:
                ticket_status['causes_overflow'] = True
                overflow_threshold_reached = True
            
            # Marquer tous les tickets ajoutÃ©s aprÃ¨s le dÃ©passement
            if overflow_threshold_reached:
                ticket_status['is_over_budget'] = True
            
            tickets_with_budget_status.append(ticket_status)
        
        # Marquer si le budget global est dÃ©passÃ©
        budget_exceeded = running_total > budget.montant
    
    # Si aucun budget n'existe, afficher tous les tickets de la pÃ©riode sÃ©lectionnÃ©e
    if not budget:
        if budget_type == 'monthly':
            period_tickets = tickets.filter(
                date_ticket__year=selected_year,
                date_ticket__month=selected_month
            ).order_by('-date_ticket', '-created_at')
        else:  # yearly
            period_tickets = tickets.filter(
                date_ticket__year=selected_year
            ).order_by('-date_ticket', '-created_at')
        
    for ticket in period_tickets:
            tickets_with_budget_status.append({
                'ticket': ticket,
                'running_total': None,
                'causes_overflow': False,
                'is_over_budget': False,
        'percentage_of_budget': 0,
        'compte': (lambda tt: (tt.accounting_entries.first().compte if tt.accounting_entries.first() else None))(ticket)
            })
    
    # Calculer les tickets de la pÃ©riode pour les statistiques
    if budget_type == 'monthly':
        period_tickets_for_stats = tickets.filter(
            date_ticket__year=selected_year,
            date_ticket__month=selected_month
        )
    else:  # yearly
        period_tickets_for_stats = tickets.filter(
            date_ticket__year=selected_year
        )
    
    period_total_amount = sum(ticket.total for ticket in period_tickets_for_stats)
    
    # --- Dashboarding par compte ---
    # Récupérer tous les tickets de la période
    period_ticket_ids = list(period_tickets_for_stats.values_list('id', flat=True))
    # Récupérer toutes les écritures comptables liées à ces tickets
    period_entries = AccountingEntry.objects.filter(ticket_id__in=period_ticket_ids)
    # Agréger par compte
    from decimal import Decimal
    dashboard_comptes = {}
    for entry in period_entries:
        compte = entry.compte or 'INCONNU'
        if compte not in dashboard_comptes:
            dashboard_comptes[compte] = {
                'compte': compte,
                'total_debit': Decimal('0'),
                'total_credit': Decimal('0'),
                'nb_ecritures': 0,
                'descriptions': set(),
            }
        # Addition propre des montants (ne JAMAIS convertir en float ici)
        if entry.debit is not None:
            try:
                dashboard_comptes[compte]['total_debit'] += entry.debit
            except Exception:
                pass
        if entry.credit is not None:
            try:
                dashboard_comptes[compte]['total_credit'] += entry.credit
            except Exception:
                pass
        dashboard_comptes[compte]['nb_ecritures'] += 1
        if entry.description:
            dashboard_comptes[compte]['descriptions'].add(entry.description)
    # Agrégation fiable côté DB pour les totaux (débit uniquement comme dans le résumé)
    from django.db.models import Sum, Value, DecimalField
    from django.db.models.functions import Coalesce
    account_totals_qs = (
        AccountingEntry.objects
        .filter(ticket_id__in=period_ticket_ids)
        .values('compte')
        .annotate(
            total_debit=Coalesce(
                Sum('debit'),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=3)),
                output_field=DecimalField(max_digits=12, decimal_places=3)
            )
        )
        .order_by('compte')
    )
    account_totals = list(account_totals_qs)

    # Distribution globale (toute la base, indépendamment de la période sélectionnée)
    global_account_totals_qs = (
        AccountingEntry.objects
        .values('compte')
        .annotate(
            total_debit=Coalesce(
                Sum('debit'),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=3)),
                output_field=DecimalField(max_digits=12, decimal_places=3)
            ),
            total_credit=Coalesce(
                Sum('credit'),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=3)),
                output_field=DecimalField(max_digits=12, decimal_places=3)
            )
        )
        .order_by('compte')
    )
    global_account_totals = list(global_account_totals_qs)

    # Conversion finale en types sérialisables (Decimal -> float)
    for c in dashboard_comptes.values():
        c['total_debit'] = float(c['total_debit'])
        c['total_credit'] = float(c['total_credit'])
        c['descriptions'] = list(c['descriptions'])
    dashboard_comptes = dict(sorted(dashboard_comptes.items(), key=lambda kv: kv[0]))
    # Convertir les sets en listes pour le template
    for c in dashboard_comptes.values():
        c['descriptions'] = list(c['descriptions'])

    context = {
        'tickets': period_tickets_for_stats,  # Tickets de la période sélectionnée
        'total_amount': total_amount,  # Total global pour les statistiques générales
        'period_total_amount': period_total_amount,  # Total de la période
        'ticket_count': tickets.count(),  # Nombre total global
        'period_ticket_count': period_tickets_for_stats.count(),  # Nombre pour la période
        'budget_info': budget_info,
        'tickets_with_budget_status': tickets_with_budget_status,
        'current_budget_type': budget_type,
        'budget_exceeded': budget_exceeded,
        'overflow_threshold_reached': overflow_threshold_reached,
        'dashboard_comptes': list(dashboard_comptes.values()),
        'account_totals': account_totals,
    'global_account_totals': global_account_totals,
    }

    return render(request, 'ocrapp/history.html', context)

@csrf_exempt
def manage_budget(request):
    """
    Vue pour gÃ©rer les budgets (crÃ©ation/modification)
    """
    from datetime import datetime
    from .models import Budget
    
    if request.method == 'POST':
        try:
            budget_type = request.POST.get('budget_type')
            montant = float(request.POST.get('montant', 0))
            current_date = datetime.now()
            
            # RÃ©cupÃ©rer les paramÃ¨tres de mois et annÃ©e ou utiliser les valeurs actuelles
            selected_year = int(request.POST.get('year', current_date.year))
            selected_month = int(request.POST.get('month', current_date.month))
            # Dynamic year range (ensure consistency with view_history)
            start_year_range = 2010
            end_year_range = current_date.year + 5
            year_range = list(range(start_year_range, end_year_range + 1))
            
            if budget_type == 'monthly':
                # CrÃ©er ou mettre Ã  jour le budget mensuel pour la pÃ©riode sÃ©lectionnÃ©e
                budget, created = Budget.objects.get_or_create(
                    type_budget='monthly',
                    annee=selected_year,
                    mois=selected_month,
                    defaults={'montant': montant}
                )
                if not created:
                    budget.montant = montant
                    budget.save()
                    
                message = f"Budget mensuel {selected_month:02d}/{selected_year} {'crÃ©Ã©' if created else 'mis Ã  jour'}: {montant} DT"
                
            else:  # yearly
                # CrÃ©er ou mettre Ã  jour le budget annuel pour l'annÃ©e sÃ©lectionnÃ©e
                budget, created = Budget.objects.get_or_create(
                    type_budget='yearly',
                    annee=selected_year,
                    defaults={'montant': montant}
                )
                if not created:
                    budget.montant = montant
                    budget.save()
                    
                message = f"Budget annuel {selected_year} {'crÃ©Ã©' if created else 'mis Ã  jour'}: {montant} DT"
            
            return JsonResponse({
                'success': True,
                'message': message,
                'budget_amount': float(budget.montant),
                'budget_type': budget_type,
                'selected_year': selected_year,
                'selected_month': selected_month,
                'year_range': year_range
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de la gestion du budget: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'MÃ©thode non autorisÃ©e'})

def download_accounting_excel(request):
    """
    Vue pour tÃ©lÃ©charger le bilan comptable en Excel
    """
    if request.method == 'POST':
        try:
            # RÃ©cupÃ©rer les donnÃ©es du formulaire
            compte = request.POST.get('compte', '606100')
            description = request.POST.get('description', 'Achat divers')
            analysis_type = request.POST.get('analysis_data', 'qwen')
            
            # RÃ©cupÃ©rer les donnÃ©es selon le type d'analyse
            ticket_id = request.POST.get('ticket_id')
            analysis_data = None

            if analysis_type == 'gemini':
                analysis_data = request.session.get('gemini_analysis')
                print(f"Utilisation des donnÃ©es Gemini: {list(analysis_data.keys()) if analysis_data else 'Aucune donnÃ©e'}")
            elif analysis_type == 'qwen':
                analysis_data = request.session.get('llm_analysis')
                print(f"Utilisation des donnÃ©es Qwen: {list(analysis_data.keys()) if analysis_data else 'Aucune donnÃ©e'}")
            elif analysis_type == 'manual' and ticket_id:
                # Charger le ticket sauvegardÃ© et fabriquer une structure compatible
                try:
                    ticket = TicketHistory.objects.get(id=int(ticket_id))
                    # Build analysis_data compatible with generate_accounting_report expectations
                    total_str = f"{ticket.total:.3f} DT" if ticket.total is not None else "0.000 DT"
                    analysis_data = {
                        'Magasin': ticket.magasin or '',
                        'Date': ticket.date_ticket.strftime('%d/%m/%Y') if ticket.date_ticket else '',
                        'Total': total_str,
                        'Articles': ticket.articles_data or []
                    }
                    print(f"Utilisation des donnÃ©es manuelles du ticket id={ticket_id}")
                except Exception as e:
                    print(f"Impossible de charger le ticket {ticket_id}: {e}")
                    analysis_data = None
            
            if not analysis_data:
                print(f"Aucune donnÃ©e {analysis_type} trouvÃ©e en session ou via ticket")
                # Try to load saved ticket as a fallback when the client provided a ticket_id
                if ticket_id:
                    try:
                        ticket = TicketHistory.objects.get(id=int(ticket_id))
                        total_str = f"{ticket.total:.3f} DT" if ticket.total is not None else "0.000 DT"
                        analysis_data = {
                            'Magasin': ticket.magasin or '',
                            'Date': ticket.date_ticket.strftime('%d/%m/%Y') if ticket.date_ticket else '',
                            'Total': total_str,
                            'Articles': ticket.articles_data or []
                        }
                        print(f"Fallback: utilisation des donnÃ©es du ticket sauvegardÃ© id={ticket_id}")
                    except Exception as e:
                        print(f"Impossible de charger le ticket {ticket_id} en fallback: {e}")

            # If still no analysis_data, create an empty structure and continue so the user can
            # generate an (empty) accounting report manually after confirming in the UI.
            if not analysis_data:
                print("Aucune donnÃ©e LLM ou ticket trouvÃ©e; gÃ©nÃ©ration manuelle avec donnÃ©es vides")
                analysis_data = {
                    'Magasin': '',
                    'Date': '',
                    'Total': '0.000 DT',
                    'Articles': []
                }
            
            # Générer le rapport comptable WITHOUT creating a new TicketHistory by default
            # We only persist an AccountingEntry if the client provided a saved ticket_id.
            report_data = generate_accounting_report(analysis_data, compte, description, save_to_db=False)
            if not report_data:
                return HttpResponse("Erreur lors de la génération du rapport", status=400)

            # If client passed a ticket_id referring to an existing saved TicketHistory,
            # create a single AccountingEntry linked to that ticket (avoid creating a new TicketHistory).
            if ticket_id:
                try:
                    existing_ticket = TicketHistory.objects.get(id=int(ticket_id))
                    # Avoid duplicate accounting entries: check if a similar entry already exists
                    try:
                        debit_decimal = Decimal(str(report_data.get('debit') or 0))
                    except Exception:
                        debit_decimal = None

                    duplicate_qs = AccountingEntry.objects.filter(
                        ticket=existing_ticket,
                        compte=compte,
                        description=description,
                    )
                    if debit_decimal is not None:
                        duplicate_qs = duplicate_qs.filter(debit=debit_decimal)

                    if not duplicate_qs.exists():
                        # parse the date for the accounting entry
                        date_ecriture = date.today()
                        date_ticket_val = report_data.get('date_ticket') or ''
                        if date_ticket_val:
                            date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]
                            for fmt in date_formats:
                                try:
                                    date_ecriture = datetime.strptime(str(date_ticket_val).split()[0], fmt).date()
                                    break
                                except:
                                    continue

                        AccountingEntry.objects.create(
                            ticket=existing_ticket,
                            date_ecriture=date_ecriture,
                            compte=compte,
                            description=description,
                            libelle_ecriture=report_data.get('libelle_ecriture', ''),
                            debit=debit_decimal if debit_decimal is not None else None,
                            credit=None
                        )
                except TicketHistory.DoesNotExist:
                    # ticket_id provided but not found - do not create a new ticket; continue without persisting
                    pass
                except Exception as e:
                    print(f"Erreur lors de la creation d'entrÃ©e comptable pour ticket_id={ticket_id}: {e}")
            
            # GÃ©nÃ©rer le fichier Excel
            excel_buffer = generate_accounting_excel(report_data)
            
            # CrÃ©er la rÃ©ponse HTTP
            response = HttpResponse(
                excel_buffer.getvalue(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="bilan_comptable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
            
        except Exception as e:
            print(f"Erreur lors de la gÃ©nÃ©ration du fichier Excel: {e}")
            return HttpResponse(f"Erreur lors de la gÃ©nÃ©ration du fichier Excel: {str(e)}", status=500)
    
    return HttpResponse("MÃ©thode non autorisÃ©e", status=405)

def upload_ticket(request):
    ocr_results = None
    llm_analysis = None
    gemini_analysis = None
    regex_analysis = None
    error = None
    form = TicketUploadForm()
    
    # Run system diagnostics on first load
    if request.method == 'GET':
        system_issues = diagnose_system()
        if system_issues:
            logger.warning("System issues detected - some features may not work properly")
    
    if request.method == 'POST':
        # Debug: afficher les donnÃ©es POST
        print("POST data:", request.POST)
        logger.info("Processing OCR request")
        
        # VÃ©rifier si on a des textes OCR dans les champs cachÃ©s
        ocr_doctr = request.POST.get('ocr_doctr')
        ocr_tesseract = request.POST.get('ocr_tesseract')
        ocr_docling = request.POST.get('ocr_docling')
        
        print("OCR fields:", {
            'doctr': bool(ocr_doctr),
            'tesseract': bool(ocr_tesseract),
            'docling': bool(ocr_docling)
        })
        
        if ocr_doctr and ocr_tesseract and ocr_docling:
            print("Using existing OCR texts")
            print(f"OCR data lengths: doctr={len(ocr_doctr)}, tesseract={len(ocr_tesseract)}, docling={len(ocr_docling)}")
            # Cas oÃ¹ on a dÃ©jÃ  les textes OCR et on veut analyser avec LLM
            ocr_results = {
                'doctr': ocr_doctr,
                'tesseract': ocr_tesseract,
                'docling': ocr_docling,
            }
            
            # RÃ©cupÃ©rer l'instance de l'image depuis la session ou la derniÃ¨re entrÃ©e
            try:
                # Essayer de rÃ©cupÃ©rer l'ID de l'image depuis la session
                image_id = request.session.get('current_image_id')
                if image_id:
                    from .models import ExtractionHistory
                    instance = ExtractionHistory.objects.get(id=image_id)
                    form = TicketUploadForm(instance=instance)
                    print(f"Image rÃ©cupÃ©rÃ©e depuis la session: {instance.image.url}")
                else:
                    # Fallback: rÃ©cupÃ©rer la derniÃ¨re image uploadÃ©e
                    from .models import ExtractionHistory
                    instance = ExtractionHistory.objects.latest('uploaded_at')
                    form = TicketUploadForm(instance=instance)
                    print(f"Image rÃ©cupÃ©rÃ©e (derniÃ¨re): {instance.image.url}")
            except Exception as e:
                print(f"Erreur lors de la rÃ©cupÃ©ration de l'image: {e}")
                # Garder le form vide si on ne peut pas rÃ©cupÃ©rer l'image
                pass
            
            # VÃ©rifier quel type d'analyse est demandÃ©
            analyze_with_gemini = request.POST.get('analyze_gemini') == '1'
            
            if analyze_with_gemini:
                print("Starting Gemini analysis...")
                gemini_analysis = analyze_three_texts_with_gemini(ocr_results)
                print(f"Gemini analysis result: {type(gemini_analysis)}")
            else:
                # Analyser avec le LLM par dÃ©faut
                print("Starting LLM analysis...")
                llm_analysis = analyze_three_texts_with_llm(ocr_results)
                print(f"LLM analysis result: {type(llm_analysis)}")
            
        else:
            print("Processing new image upload")
            # Cas normal avec upload d'image
            form = TicketUploadForm(request.POST, request.FILES)
            if form.is_valid():
                instance = form.save()
                image_path = instance.image.path
                
                # Sauvegarder l'ID de l'image en session pour les analyses ultÃ©rieures
                request.session['current_image_id'] = instance.id
                print(f"Image ID sauvegardÃ© en session: {instance.id}")

                # VÃ©rifier quel bouton a Ã©tÃ© cliquÃ©
                analyze_all = request.POST.get('ocr_all') == '1'
                analyze_with_llm = request.POST.get('analyze_llm') == '1'
                analyze_with_gemini = request.POST.get('analyze_gemini') == '1'
                analyze_all_models = request.POST.get('analyze_all_models') == '1'
                analyze_with_regex = request.POST.get('analyze_regex') == '1'
                
                print("Analysis flags:", {
                    'analyze_all': analyze_all,
                    'analyze_all_models': analyze_all_models,
                    'analyze_with_llm': analyze_with_llm,
                    'analyze_with_gemini': analyze_with_gemini,
                    'analyze_with_regex': analyze_with_regex
                })

                # Si on clique sur "Analyser le ticket (OCR + Gemini + Qwen)"
                if analyze_all and analyze_all_models:
                    print("Complete analysis: OCR + Gemini + LLM (combined)")
                    # Lancer les 3 OCR
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    # Puis analyser avec Gemini et LLM
                    if ocr_results:
                        gemini_analysis = analyze_three_texts_with_gemini(ocr_results)
                        try:
                            print(f"Gemini analysis produced: type={type(gemini_analysis)}, keys={list(gemini_analysis.keys()) if isinstance(gemini_analysis, dict) else 'N/A'}")
                        except Exception:
                            pass
                        llm_analysis = analyze_three_texts_with_llm(ocr_results)
                        try:
                            print(f"LLM analysis produced: type={type(llm_analysis)}, keys={list(llm_analysis.keys()) if isinstance(llm_analysis, dict) else 'N/A'}")
                        except Exception:
                            pass
                        # Ajouter aussi l'analyse regex pour comparaison
                        texte_combine = f"{ocr_results['doctr']}\n{ocr_results['tesseract']}\n{ocr_results['docling']}"
                        regex_analysis = extraire_elements_avec_regex(texte_combine)

                # Si on clique sur "Extraction OCR uniquement" (bouton bleu)
                # Ne lancer l'extraction seule que si aucun LLM (Qwen/Gemini) n'est demandé
                elif analyze_all and not analyze_with_llm and not analyze_with_gemini:
                    print("Extracting OCR texts only")
                    # Lancer les 3 OCR
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                
                # Si on clique sur "OCR + Analyse Qwen3-30B" (bouton vert)
                # Ne lancer Qwen que si Gemini n'a pas été demandé (Gemini prend la priorité)
                elif analyze_with_llm and analyze_all and not analyze_with_gemini and not analyze_all_models:
                    print("Extracting OCR and analyzing with LLM")
                    # Extraire les OCR puis analyser avec le LLM
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    
                    # Puis analyser avec le LLM
                    if ocr_results:
                        llm_analysis = analyze_three_texts_with_llm(ocr_results)
                        # Debug: log LLM analysis result shape for troubleshooting
                        try:
                            print(f"LLM analysis produced: type={type(llm_analysis)}, keys={list(llm_analysis.keys()) if isinstance(llm_analysis, dict) else 'N/A'}")
                            logger.info(f"LLM analysis produced keys: {list(llm_analysis.keys()) if isinstance(llm_analysis, dict) else 'non-dict result'}")
                        except Exception as e:
                            print(f"Erreur en loggant le résultat LLM: {e}")
                
                # Si on clique sur "OCR + Analyse Google Gemini" (bouton violet)
                # Gemini prend la priorité même si l'autre flag est présent
                elif analyze_with_gemini and analyze_all and not analyze_all_models:
                    print("Extracting OCR and analyzing with Gemini")
                    # Extraire les OCR puis analyser avec Gemini
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    
                    # Puis analyser avec Gemini
                    if ocr_results:
                        gemini_analysis = analyze_three_texts_with_gemini(ocr_results)
                        # Debug: log Gemini analysis result shape for troubleshooting
                        try:
                            print(f"Gemini analysis produced: type={type(gemini_analysis)}, keys={list(gemini_analysis.keys()) if isinstance(gemini_analysis, dict) else 'N/A'}")
                            logger.info(f"Gemini analysis produced keys: {list(gemini_analysis.keys()) if isinstance(gemini_analysis, dict) else 'non-dict result'}")
                        except Exception as e:
                            print(f"Erreur en loggant le résultat Gemini: {e}")
                        # Ajouter aussi l'analyse regex pour comparaison
                        texte_combine = f"{ocr_results['doctr']}\n{ocr_results['tesseract']}\n{ocr_results['docling']}"
                        regex_analysis = extraire_elements_avec_regex(texte_combine)
                
                # Si on clique sur "Analyse Regex (Rapide)" (bouton jaune)
                elif analyze_with_regex and not analyze_all:
                    print("Extracting OCR and analyzing with Regex")
                    # Extraire les OCR puis analyser avec regex
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    
                    # Puis analyser avec regex
                    if ocr_results:
                        texte_combine = f"{ocr_results['doctr']}\n{ocr_results['tesseract']}\n{ocr_results['docling']}"
                        regex_analysis = extraire_elements_avec_regex(texte_combine)
                        print(f"Regex analysis result: {regex_analysis}")
                
                # Si on clique sur "Analyse ComplÃ¨te" (bouton info) - Qwen uniquement si Gemini non demandÃ©
                elif analyze_all and analyze_with_llm and not analyze_with_gemini and not analyze_all_models:
                    print("Complete analysis: OCR + LLM")
                # Si on clique sur "Analyser le ticket (OCR + Gemini + Qwen)"
                elif analyze_all and analyze_all_models:
                    print("Complete analysis: OCR + Gemini + LLM (combined)")
                    # Lancer les 3 OCR
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    # Puis analyser avec Gemini et LLM
                    if ocr_results:
                        gemini_analysis = analyze_three_texts_with_gemini(ocr_results)
                        try:
                            print(f"Gemini analysis produced: type={type(gemini_analysis)}, keys={list(gemini_analysis.keys()) if isinstance(gemini_analysis, dict) else 'N/A'}")
                        except Exception:
                            pass
                        llm_analysis = analyze_three_texts_with_llm(ocr_results)
                        try:
                            print(f"LLM analysis produced: type={type(llm_analysis)}, keys={list(llm_analysis.keys()) if isinstance(llm_analysis, dict) else 'N/A'}")
                        except Exception:
                            pass
                        # Ajouter aussi l'analyse regex pour comparaison
                        texte_combine = f"{ocr_results['doctr']}\n{ocr_results['tesseract']}\n{ocr_results['docling']}"
                        regex_analysis = extraire_elements_avec_regex(texte_combine)
                    # Lancer les 3 OCR
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    
                    # Puis analyser avec le LLM
                    if ocr_results:
                        llm_analysis = analyze_three_texts_with_llm(ocr_results)

    # Convertir les donnÃ©es LLM en JSON pour le formulaire
    llm_analysis_json = None
    if llm_analysis and not isinstance(llm_analysis, str):
        try:
            # Filtrer les donnÃ©es sensibles et s'assurer que les valeurs sont sÃ©rialisables
            clean_llm_data = {}
            for key, value in llm_analysis.items():
                if key not in ['texte_fusionne', 'raw_response']:
                    # S'assurer que la valeur est sÃ©rialisable
                    if isinstance(value, (str, int, float, list, dict, bool)) or value is None:
                        clean_llm_data[key] = value
            
            # Sauvegarder en session pour le tÃ©lÃ©chargement PDF
            request.session['llm_analysis'] = clean_llm_data
            
            llm_analysis_json = json.dumps(clean_llm_data, ensure_ascii=False, separators=(',', ':'))
            print(f"JSON gÃ©nÃ©rÃ©: {llm_analysis_json[:200]}...")
        except Exception as e:
            print(f"Erreur lors de la gÃ©nÃ©ration du JSON: {e}")
            llm_analysis_json = None
    
    # Sauvegarder gemini_analysis en session si disponible
    if gemini_analysis and not isinstance(gemini_analysis, str):
        try:
            # Filtrer les donnÃ©es sensibles et s'assurer que les valeurs sont sÃ©rialisables
            clean_gemini_data = {}
            for key, value in gemini_analysis.items():
                if key not in ['texte_fusionne', 'raw_response']:
                    # S'assurer que la valeur est sÃ©rialisable
                    if isinstance(value, (str, int, float, list, dict, bool)) or value is None:
                        clean_gemini_data[key] = value
            
            # Sauvegarder en session pour le tÃ©lÃ©chargement PDF
            request.session['gemini_analysis'] = clean_gemini_data
            print(f"DonnÃ©es Gemini sauvegardÃ©es en session: {list(clean_gemini_data.keys())}")
        except Exception as e:
            print(f"Erreur lors de la sauvegarde Gemini en session: {e}")
    
    # Ensure template can safely access keys like .Magasin/.Total without raising
    # VariableDoesNotExist when analysis variables are None. Convert to empty
    # dicts so template lookups return empty values instead of failing.
    if llm_analysis is None:
        llm_analysis = {}
    if gemini_analysis is None:
        gemini_analysis = {}
    if regex_analysis is None:
        regex_analysis = {}

    # Compute safe accounting-related context values so the template doesn't
    # attempt to resolve nested variables that may be missing (which was
    # causing VariableDoesNotExist exceptions when using default:other_var
    # syntax in the template).
    accounting_analysis_choice = ''
    accounting_llm_data = ''
    try:
        # Prefer manual saved values first, then Gemini, then LLM
        try:
            inst_magasin = getattr(form.instance, 'magasin', None)
            inst_total = getattr(form.instance, 'total', None)
            if inst_magasin and inst_total:
                accounting_analysis_choice = 'manual'
                accounting_llm_data = str(inst_magasin)
            else:
                if isinstance(gemini_analysis, dict) and gemini_analysis.get('Magasin') and gemini_analysis.get('Total'):
                    accounting_analysis_choice = 'gemini'
                    accounting_llm_data = str(gemini_analysis.get('Magasin') or '')
                elif isinstance(llm_analysis, dict) and llm_analysis.get('Magasin') and llm_analysis.get('Total'):
                    accounting_analysis_choice = 'qwen'
                    accounting_llm_data = str(llm_analysis.get('Magasin') or '')
        except Exception:
            # If anything goes wrong accessing form.instance, fall back to LLM/Gemini
            if isinstance(gemini_analysis, dict) and gemini_analysis.get('Magasin') and gemini_analysis.get('Total'):
                accounting_analysis_choice = 'gemini'
                accounting_llm_data = str(gemini_analysis.get('Magasin') or '')
            elif isinstance(llm_analysis, dict) and llm_analysis.get('Magasin') and llm_analysis.get('Total'):
                accounting_analysis_choice = 'qwen'
                accounting_llm_data = str(llm_analysis.get('Magasin') or '')
    except Exception:
        # Be defensive: leave defaults as empty strings on any unexpected shape
        accounting_analysis_choice = ''
        accounting_llm_data = ''

    return render(request, 'ocrapp/upload_ticket.html', {
        'form': form,
        'ocr_results': ocr_results,
        'llm_analysis': llm_analysis,
        'gemini_analysis': gemini_analysis,
        'regex_analysis': regex_analysis,
        'llm_analysis_json': llm_analysis_json,
        'error': error,
        'accounting_analysis_choice': accounting_analysis_choice,
        'accounting_llm_data': accounting_llm_data,
    })

@csrf_exempt
def filter_accounting_data(request):
    """
    Filtre les donnÃ©es comptables par date pour l'affichage AJAX
    """
    if request.method == 'POST':
        try:
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            
            # Construire la requÃªte de base
            entries_query = AccountingEntry.objects.all()

            # Exclure les deux lignes spÃ©cifiques demandÃ©es (paiements AZIZA 4.090 DT)
            try:
                from datetime import date as _date
                entries_query = entries_query.exclude(
                    compte='531200',
                    description='Paiement ticket de caisse',
                    libelle_ecriture='Paiement ticket - AZIZA - 80102080 - 4.090 DT',
                    date_ecriture__in=[_date(2025, 2, 1), _date(2025, 8, 8)]
                )
            except Exception:
                pass
            
            # Appliquer les filtres de date si fournis
            if start_date:
                try:
                    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                    entries_query = entries_query.filter(date_ecriture__gte=start_date_obj)
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Format de date de dÃ©but invalide'})
            
            if end_date:
                try:
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                    entries_query = entries_query.filter(date_ecriture__lte=end_date_obj)
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Format de date de fin invalide'})
            
            # RÃ©cupÃ©rer les entrÃ©es filtrÃ©es
            entries = entries_query.order_by('-date_ecriture')
            
            # PrÃ©parer les donnÃ©es pour JSON
            entries_data = []
            for entry in entries:
                entries_data.append({
                    'date_ecriture': entry.date_ecriture.strftime('%Y-%m-%d'),
                    'compte': entry.compte,
                    'description': entry.description,
                    'libelle_ecriture': entry.libelle_ecriture,
                    'debit': float(entry.debit),
                    'credit': float(entry.credit) if entry.credit else 0.0
                })
            
            # Calculer les totaux par compte
            account_totals = {}
            total_debit = Decimal('0')
            total_credit = Decimal('0')
            
            for entry in entries:
                account = entry.compte
                if account not in account_totals:
                    account_totals[account] = {'debit': Decimal('0'), 'credit': Decimal('0')}
                
                account_totals[account]['debit'] += entry.debit
                if entry.credit:
                    account_totals[account]['credit'] += entry.credit
                
                total_debit += entry.debit
                if entry.credit:
                    total_credit += entry.credit
            
            # Convertir les totaux en float pour JSON
            account_totals_json = {}
            for account, totals in account_totals.items():
                account_totals_json[account] = {
                    'debit': float(totals['debit']),
                    'credit': float(totals['credit'])
                }
            
            # RÃ©sumÃ© gÃ©nÃ©ral
            summary = {
                'total_entries': len(entries_data),
                'total_debit': float(total_debit),
                'total_credit': float(total_credit),
                'total_accounts': len(account_totals),
                'period': {
                    'start': start_date if start_date else 'DÃ©but',
                    'end': end_date if end_date else 'Fin'
                }
            }
            
            return JsonResponse({
                'success': True,
                'entries': entries_data,
                'account_totals': account_totals_json,
                'summary': summary
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du filtrage des donnÃ©es comptables: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Erreur serveur: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'MÃ©thode non autorisÃ©e'})


try:
    import google.generativeai as genai
    genai.configure(api_key=os.environ.get("GOOGLE_API_KEY"))
except Exception as _genai_import_err:
    genai = None
    logger.warning("Google Generative AI client not available - Gemini integration disabled: %s", _genai_import_err)

def verifier_par_gemini(qwen_data, ocr_texts_combined):
    """
    Utilise Gemini 2.5 Flash pour valider les informations extraites par Qwen
    """
    prompt = f"""
Tu es un assistant spÃ©cialisÃ© dans la validation de donnÃ©es extraites de tickets de caisse.

Voici un texte brut issu de plusieurs OCR :
-------------------------------
{ocr_texts_combined}
-------------------------------

Et voici les donnÃ©es extraites par un autre modÃ¨le (Qwen) :

{json.dumps(qwen_data, indent=2, ensure_ascii=False)}

Analyse attentivement le texte OCR et dis-moi si les informations suivantes semblent cohÃ©rentes avec le contenu :
- Le nom du magasin
- La date et heure
- Le numÃ©ro du ticket
- Les articles (noms et prix)
- Le montant total

RÃ‰PONDS uniquement avec un JSON structurÃ© comme ceci :

{{
  "verdict": "valide" ou "invalide",
  "problemes_detectes": ["description 1", "description 2", ...],
  "suggestions": ["correction 1", "correction 2", ...]
}}

IMPORTANT : ne donne que ce JSON.
    """

    # If the Google Generative client isn't available, return a safe response
    if genai is None:
        logger.info("verifier_par_gemini called but Gemini client is unavailable")
        return {
            "verdict": "disabled",
            "problemes_detectes": ["Gemini integration unavailable on server"],
            "suggestions": []
        }

    try:
        model = genai.GenerativeModel("gemini-2.5-flash")  # Ou "gemini-2.5-flash" si dispo
        response = model.generate_content(prompt)
        return clean_json_response(response.text)
    except Exception as e:
        logger.exception("Erreur avec Gemini: %s", e)
        return {
            "verdict": "erreur",
            "problemes_detectes": [f"Erreur Gemini: {str(e)}"],
            "suggestions": []
        }

@csrf_exempt
def get_ticket_details(request, ticket_id):
    """
    Vue pour rÃ©cupÃ©rer les dÃ©tails d'un ticket pour modification
    """
    if request.method == 'GET':
        try:
            ticket = Ticket.objects.get(id=ticket_id)
            return JsonResponse({
                'success': True,
                'ticket': {
                    'id': ticket.id,
                    'magasin': ticket.magasin,
                    'numero_ticket': ticket.numero_ticket,
                    'date_ticket': ticket.date_ticket.strftime('%Y-%m-%d'),
                    'total': float(ticket.total),
                    'articles_data': ticket.articles_data or []
                }
            })
        except Ticket.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Ticket non trouvÃ©'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de la rÃ©cupÃ©ration du ticket: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'MÃ©thode non autorisÃ©e'})

@csrf_exempt
def update_ticket(request, ticket_id):
    """
    Vue pour mettre Ã  jour un ticket
    """
    if request.method == 'POST':
        try:
            ticket = Ticket.objects.get(id=ticket_id)
            
            # RÃ©cupÃ©rer les donnÃ©es du formulaire
            magasin = request.POST.get('magasin')
            numero_ticket = request.POST.get('numero_ticket')
            date_ticket = request.POST.get('date_ticket')
            total = request.POST.get('total')
            articles_data = request.POST.get('articles_data')
            
            # Mettre Ã  jour les champs
            if magasin:
                ticket.magasin = magasin
            if numero_ticket:
                ticket.numero_ticket = numero_ticket
            if date_ticket:
                ticket.date_ticket = datetime.strptime(date_ticket, '%Y-%m-%d').date()
            if total:
                ticket.total = Decimal(total)
            if articles_data:
                import json
                ticket.articles_data = json.loads(articles_data)
            
            ticket.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket mis Ã  jour avec succÃ¨s',
                'ticket': {
                    'id': ticket.id,
                    'magasin': ticket.magasin,
                    'numero_ticket': ticket.numero_ticket,
                    'date_ticket': ticket.date_ticket.strftime('%Y-%m-%d'),
                    'total': float(ticket.total),
                    'articles_data': ticket.articles_data or []
                }
            })
            
        except Ticket.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Ticket non trouvÃ©'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de la mise Ã  jour du ticket: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'MÃ©thode non autorisÃ©e'})
@csrf_exempt
def save_ticket_analysis(request):
    """
    Vue pour sauvegarder les donnÃ©es modifiÃ©es d'un ticket aprÃ¨s analyse LLM
    """
    if request.method == 'POST':
        try:
            # RÃ©cupÃ©rer les donnÃ©es du formulaire
            magasin = request.POST.get('magasin')
            date_str = request.POST.get('date')
            numero_ticket = request.POST.get('numero')
            total_str = request.POST.get('total')
            tva_rate_str = request.POST.get('tva_rate')
            tva_amount_str = request.POST.get('tva_amount')
            articles_json = request.POST.get('articles')
            analysis_type = request.POST.get('analysis_type', 'qwen')  # qwen ou gemini
            
            # Validation des donnÃ©es obligatoires
            if not all([magasin, date_str, total_str]):
                return JsonResponse({
                    'success': False,
                    'error': 'Les champs magasin, date et total sont obligatoires'
                })

            # Fonction pour nettoyer les valeurs numÃšriques
            def clean_numeric_value(value):
                """Extrait la valeur numÃšrique d'une chaÃ¯ne en supprimant les unitÃšs et caractÃ¨res non numÃšriques"""
                if not value or value == 'None':
                    return '0'

                import re
                # Convertir en chaÃ¯ne et nettoyer
                value_str = str(value).strip()

                # Remplacer les virgules par des points
                value_str = value_str.replace(',', '.')

                # Extraire uniquement les chiffres, points et signes moins
                numeric_match = re.search(r'[-]?\d*\.?\d+', value_str)
                if numeric_match:
                    return numeric_match.group(0)
                else:
                    return '0'

            # Nettoyage et validation des donnÃšes
            try:
                # Nettoyage des chaÃ¯nes numÃšriques
                total_str = clean_numeric_value(total_str)
                tva_rate_str = clean_numeric_value(tva_rate_str)
                tva_amount_str = clean_numeric_value(tva_amount_str)

                # Validation et conversion de la date
                date_ticket = datetime.strptime(date_str, '%Y-%m-%d').date()

                # Validation et conversion des montants
                if not total_str or total_str == 'None':
                    return JsonResponse({
                        'success': False,
                        'error': 'Le montant total est obligatoire'
                    })
                
                total = Decimal(total_str)
                tva_rate = Decimal(tva_rate_str) if tva_rate_str and tva_rate_str != '0' else Decimal('0')
                tva_amount = Decimal(tva_amount_str) if tva_amount_str and tva_amount_str != '0' else Decimal('0')
                
                # VÃ©rification des valeurs nÃgatives
                if total < 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'Le montant total ne peut pas Ãªtre nÃgatif'
                    })
                    
            except (ValueError, InvalidOperation, TypeError) as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Erreur de format dans les donnÃ©es: total="{total_str}", tva_rate="{tva_rate_str}", tva_amount="{tva_amount_str}". Erreur: {str(e)}'
                })
            
            # Parser les articles
            articles_data = []
            if articles_json:
                try:
                    import json
                    articles_data = json.loads(articles_json)
                except json.JSONDecodeError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Format JSON invalide pour les articles'
                    })
            
            # Récupérer les compte/description envoyés depuis le front (valeurs modifiables par l'utilisateur)
            compte = request.POST.get('compte', '606100')
            description = request.POST.get('description', 'Ticket de caisse')

            # Créer ou mettre à jour le ticket
            ticket = TicketHistory.objects.filter(
                magasin=magasin,
                date_ticket=date_ticket,
                numero_ticket=numero_ticket or ''
            ).first()
            created = False
            if not ticket:
                ticket = TicketHistory.objects.create(
                    magasin=magasin,
                    date_ticket=date_ticket,
                    numero_ticket=numero_ticket or '',
                    total=total,
                    articles_data=articles_data,
                    tva_rate=tva_rate,
                    tva_amount=tva_amount,
                    llm_analysis={'compte': compte, 'description': description}
                )
                created = True
            else:
                # Mettre à jour le ticket existant
                ticket.total = total
                ticket.articles_data = articles_data
                ticket.tva_rate = tva_rate
                ticket.tva_amount = tva_amount
                # Mettre à jour aussi l'analyse LLM persistée avec les valeurs de compte/description
                try:
                    existing_llm = ticket.llm_analysis or {}
                    if not isinstance(existing_llm, dict):
                        existing_llm = {}
                    existing_llm['compte'] = compte
                    existing_llm['description'] = description
                    ticket.llm_analysis = existing_llm
                except Exception:
                    ticket.llm_analysis = {'compte': compte, 'description': description}
                ticket.save()

            # Créer les entrées comptables
            if not created:
                # Si le ticket existait déjà, supprimer ses anciennes entrées comptables
                AccountingEntry.objects.filter(ticket=ticket).delete()

            # Créer les nouvelles entrées comptables spécifiques aux tickets de caisse
            # Entrée au débit (Achat - Ticket de caisse)
            AccountingEntry.objects.create(
                ticket=ticket,
                date_ecriture=date_ticket,
                compte=compte,
                description=description,
                libelle_ecriture=f'Ticket de caisse - {magasin} - {numero_ticket or "N/A"} - {total} DT',
                debit=total,
                credit=Decimal('0')
            )
            
            # NOTE: Removed creation of the corresponding credit entry (paiement)
            # per request: tickets must only be recorded in the debit side.
            
            # Retourner les données mises à jour pour le bilan comptable
            return JsonResponse({
                'success': True,
                'message': f'Ticket de caisse {"créé" if created else "mis à jour"} avec succès et ajouté au bilan comptable',
                'ticket_id': ticket.id,
                'created': created,
                'updated_data': {
                    'date': date_ticket.strftime('%d/%m/%Y %H:%M'),
                    'magasin': magasin,
                    'total': float(total),
                    'libelle': f'Ticket de caisse - {magasin} - {numero_ticket or "N/A"} - {total} DT',
                    'compte': compte,
                    'description': description
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de la sauvegarde: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'MÃ©thode non autorisÃ©e'})
